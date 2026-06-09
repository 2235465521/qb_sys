"""
standards.services — 标准业务逻辑层

功能：
  - 特殊标准号 clean_id 生成（严格区分前缀）
  - 规范性引用 Excel 解析（模块二）
  - 国标引用热度统计
  - ZIP 打包异步任务
"""

import re
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.db.models import F, Q

from .models import Standard, NormativeReference


# ============================================================
# 特殊标准号处理
# ⚠️ 关键规则：严格保留省份简称前缀和少数民族语言代码
# ============================================================

# 已知特殊前缀模式（少数民族语言代码、省份简称等）
# 必须放在通用前缀匹配之前
SPECIAL_PREFIX_PATTERNS = [
    # 少数民族语言代码：如 DBY（彝文）、DBZ（藏文）等
    r'^(DB[A-Z]{1,2})',
    # 省份简称+标准类型：如 新Q/T、藏DB、蒙Q/T
    r'^([新藏蒙琼黔陕甘皖滇闽鄂赣豫湘苏浙粤川吉辽冀津沪京渝宁桂内][A-Z])',
    # 带斜杠的特殊编号
    r'^([A-Z]{2,4}/[A-Z])',
]

def generate_clean_id(standard_no: str) -> str:
    """
    生成标准化 clean_id，用于检索和去重。

    ⚠️ 核心规则：
    1. 保留所有特殊前缀（"新Q"、"DBY"等）
    2. 不同前缀 = 不同标准，绝不合并
    3. 仅移除多余空格，统一大写

    Args:
        standard_no: 原始标准号（如 "新Q/T 123-2026"、"DBY/T 001-2025"）

    Returns:
        标准化 clean_id（如 "新Q/T123-2026"、"DBY/T001-2025"）
    """
    if not standard_no:
        return ''

    # 去除首尾空格，统一大写字母部分
    clean = standard_no.strip()

    # 移除内部多余空格（但保留前缀中的汉字）
    # 只移除数字/字母之间的空格（版本号前的空格）
    clean = re.sub(r'(?<=[A-Z0-9])\s+(?=[0-9])', '', clean)

    # 统一连字符
    clean = re.sub(r'[—–]', '-', clean)

    return clean


def validate_standard_no(standard_no: str) -> bool:
    """
    验证标准号格式是否合法。
    接受所有已知前缀（包含特殊前缀），不能因校验丢失数据。
    """
    if not standard_no or len(standard_no.strip()) < 2:
        return False
    # 只要不是空值，都接受（宽松校验，不能丢数据）
    return True


# ============================================================
# 规范化查询服务 (Exact / Fuzzy Search with Fault Tolerance)
# ============================================================

def parse_date_param(param_start, param_end):
    from datetime import date, datetime
    try:
        if len(param_start) == 4:
            start_date = date(int(param_start), 1, 1)
        else:
            start_date = datetime.strptime(param_start, '%Y-%m-%d').date()

        if len(param_end) == 4:
            end_date = date(int(param_end), 12, 31)
        else:
            end_date = datetime.strptime(param_end, '%Y-%m-%d').date()

        return start_date, end_date
    except Exception:
        return None, None

def search_standards_service(params):
    """
    分离自 StandardListView 的标准检索核心逻辑。
    支持：地域、状态、时间过滤，并且针对关键词做了基于 clean_id 的多级容错与精确查询。
    """
    qs = Standard.objects.select_related('company').prefetch_related('normative_references')

    if params.get('type'):
        qs = qs.filter(type=params['type'])

    if params.get('is_parsed') in ('true', 'false'):
        qs = qs.filter(is_parsed=params['is_parsed'] == 'true')

    if params.get('company_id'):
        qs = qs.filter(company_id=params['company_id'])

    # 1. 地域筛选
    province_id = params.get('province_id')
    city_id = params.get('city_id')
    district_id = params.get('district_id')
    if province_id:
        qs = qs.filter(company__province_id=province_id)
    if city_id:
        qs = qs.filter(company__city_id=city_id)
    if district_id:
        qs = qs.filter(company__district_id=district_id)

    # 2. 时间维度筛选
    pub_start = params.get('pub_start')
    pub_end = params.get('pub_end')
    imp_start = params.get('imp_start')
    imp_end = params.get('imp_end')

    if (pub_start and pub_end) or (imp_start and imp_end):
        time_filters = Q()
        if pub_start and pub_end:
            pub_s, pub_e = parse_date_param(pub_start, pub_end)
            if pub_s and pub_e:
                time_filters &= Q(publish_date__range=(pub_s, pub_e))
        if imp_start and imp_end:
            imp_s, imp_e = parse_date_param(imp_start, imp_end)
            if imp_s and imp_e:
                time_filters &= Q(implement_date__range=(imp_s, imp_e))
        if time_filters:
            qs = qs.filter(time_filters)

    # 3. 解析状态细化筛选
    parse_status = params.get('parse_status')
    if parse_status:
        statuses = parse_status.split(',') if isinstance(parse_status, str) else parse_status
        q_status = Q()
        if 'pending_reference' in statuses:
            q_status |= Q(is_parsed='unparsed')
        if 'pending_indicator' in statuses:
            q_status |= Q(is_parsed='references_parsed')
        if q_status:
            qs = qs.filter(q_status)

    # 4. 关键词容错与精确查询
    keyword = params.get('keyword')
    if keyword:
        keyword = keyword.strip()
        search_mode = params.get('search_mode', 'title')
        exact_match = params.get('exact_match') == 'true'

        if search_mode == 'full_text':
            from standards.models import StandardContent
            # PDF正文检索暂不使用 clean_id 容错，直接使用原生
            matching_std_ids = StandardContent.objects.filter(
                content__icontains=keyword
            ).values_list('standard_id', flat=True).distinct()
            qs = qs.filter(id__in=matching_std_ids)
        else:
            # 使用 clean_id 容错处理
            kw_clean = generate_clean_id(keyword)
            if exact_match:
                # 精确匹配：在清洗后的 clean_id 严格相等，或者 title 完全匹配
                qs = qs.filter(Q(clean_id=kw_clean) | Q(title__iexact=keyword))
            else:
                # 模糊匹配：在清洗后的 clean_id 做包含查询，或者 title 模糊匹配
                qs = qs.filter(Q(clean_id__icontains=kw_clean) | Q(title__icontains=keyword))

    ordering = params.get('ordering')
    if ordering:
        allowed_orderings = {
            'publish_date': 'publish_date',
            '-publish_date': '-publish_date',
            'implement_date': 'implement_date',
            '-implement_date': '-implement_date',
            'standard_no': 'standard_no',
            '-standard_no': '-standard_no',
            'title': 'title',
            '-title': '-title',
            'company_name': 'company__name',
            '-company_name': '-company__name',
            'created_at': 'created_at',
            '-created_at': '-created_at',
        }
        db_ordering = allowed_orderings.get(ordering)
        if db_ordering:
            return qs.order_by(db_ordering)

    return qs.order_by('-created_at')


# ============================================================
# 规范性引用 Excel 解析（模块二核心）
# ============================================================

CITATION_EXCEL_REQUIRED_COLS = ['企标号', '被引用标准号']


def parse_normative_reference_excel(file_obj) -> dict:
    """
    解析规范性引用 Excel 文件（支持原 2 列格式与新 10 列批量批次格式）
    """
    import pandas as pd
    import uuid
    from django.db import transaction
    from django.db.models import F
    from companies.models import Company
    from standards.models import Standard, NormativeReference

    result = {
        'parsed_standards': 0, 
        'citations_added': 0, 
        'companies_created': 0, 
        'standards_created': 0,
        'pdf_aligned_count': 0,
        'errors': []
    }

    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        result['errors'].append(f'文件解析失败: {str(e)}')
        return result

    if df.empty:
        result['errors'].append('文件内容为空')
        return result

    # 1. 检测是否是 10 列批次导入格式 (通过包含的核心列 '企标号', '公司名称', '企标中引用的标准号')
    if '企标号' in df.columns and '公司名称' in df.columns and '企标中引用的标准号' in df.columns:
        # 重命名“最新标准号”列变体以支持容错
        if '最新标准号' not in df.columns:
            for col in df.columns:
                if '最新标准号' in str(col) or '最新被引用标准号' in str(col):
                    df.rename(columns={col: '最新标准号'}, inplace=True)
                    break
        with transaction.atomic():
            for idx, row in df.iterrows():
                row_idx = idx + 2  # Excel 行号 1-indexed (加上表头)
                
                company_name = str(row['公司名称']).strip() if pd.notna(row['公司名称']) else ''
                std_no = str(row['企标号']).strip() if pd.notna(row['企标号']) else ''
                std_name = str(row['企标名']).strip() if '企标名' in df.columns and pd.notna(row['企标名']) else ''
                
                # 匹配被引用标准号（优先完整标准号，其次中引标准号）
                cited_no = None
                if '发布时引用的完整标准号' in df.columns and pd.notna(row['发布时引用的完整标准号']):
                    cited_no = str(row['发布时引用的完整标准号']).strip()
                elif pd.notna(row['企标中引用的标准号']):
                    cited_no = str(row['企标中引用的标准号']).strip()

                latest_no = ''
                if '最新标准号' in df.columns and pd.notna(row['最新标准号']):
                    latest_no = str(row['最新标准号']).strip()
                else:
                    latest_no = cited_no or ''
                    
                if not company_name or not std_no or not cited_no:
                    continue
                    
                try:
                    # 1. 自动对齐/建档企业
                    company, created_co = Company.objects.get_or_create(
                        name=company_name,
                        defaults={
                            'credit_code': f'TEMP_{uuid.uuid4().hex[:14].upper()}',
                            'status': 'active'
                        }
                    )
                    if created_co:
                        result['companies_created'] += 1
                        
                    # 2. 自动对齐/建档企标
                    clean_id = generate_clean_id(std_no)
                    standard, created_std = Standard.objects.get_or_create(
                        standard_no=std_no,
                        defaults={
                            'clean_id': clean_id,
                            'title': std_name or '批量入库标准',
                            'company': company,
                            'type': 'enterprise',
                            'is_parsed': True
                        }
                    )
                    if created_std:
                        result['standards_created'] += 1
                    else:
                        if not standard.is_parsed:
                            standard.is_parsed = True
                            standard.save(update_fields=['is_parsed'])
                            result['parsed_standards'] += 1
                            
                    # 3. 关联或建档被引用的国标（统计引用热度）
                    cited_std = Standard.objects.filter(standard_no=cited_no, type='national').first()
                    
                    # 4. 创建规范性引用关系记录
                    ref, created_ref = NormativeReference.objects.get_or_create(
                        source_standard=standard,
                        cited_standard_no=cited_no,
                        defaults={
                            'cited_standard': cited_std,
                            'latest_standard_no': latest_no
                        }
                    )
                    if not created_ref and latest_no and ref.latest_standard_no != latest_no:
                        ref.latest_standard_no = latest_no
                        ref.save(update_fields=['latest_standard_no'])

                    if created_ref:
                        result['citations_added'] += 1
                        # 累加国标引用热度计数
                        if cited_std:
                            Standard.objects.filter(pk=cited_std.pk).update(citation_count=F('citation_count') + 1)
                            
                except Exception as e:
                    result['errors'].append(f"第 {row_idx} 行解析失败: {str(e)}")
                    
        # 5. 后置自动扫盘匹配一次 PDF 相对路径
        try:
            scan_res = scan_and_align_pdf_assets()
            result['pdf_aligned_count'] = scan_res.get('matched_count', 0)
        except Exception as scan_err:
            result['errors'].append(f"关联扫盘对齐 PDF 失败: {str(scan_err)}")
            
        return result

    else:
        # 2. 原来的 2 列格式兼容导入
        parsed_standard_nos = set()
        citation_pairs = []  # [(企标号, 被引用标准号), ...]
        
        for idx, row in df.iterrows():
            row_idx = idx + 2
            if len(row) < 2:
                continue
            source_no = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            cited_no = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
            
            if not source_no:
                continue
            parsed_standard_nos.add(source_no)
            if cited_no:
                citation_pairs.append((source_no, cited_no))

        with transaction.atomic():
            updated = Standard.objects.filter(
                standard_no__in=parsed_standard_nos,
                type='enterprise'
            ).update(is_parsed=True)
            result['parsed_standards'] = updated

            if updated < len(parsed_standard_nos):
                not_found = parsed_standard_nos - set(
                    Standard.objects.filter(
                        standard_no__in=parsed_standard_nos
                    ).values_list('standard_no', flat=True)
                )
                for no in not_found:
                    result['errors'].append(f'企标未找到: {no}（可能未入库）')

            source_map = {
                s.standard_no: s
                for s in Standard.objects.filter(standard_no__in=parsed_standard_nos)
            }

            cited_nos = {cited_no for _, cited_no in citation_pairs}
            cited_map = {
                s.standard_no: s
                for s in Standard.objects.filter(
                    standard_no__in=cited_nos,
                    type='national'
                )
            }

            citation_count_delta = {}

            for source_no, cited_no in citation_pairs:
                source = source_map.get(source_no)
                if not source:
                    continue

                cited = cited_map.get(cited_no)

                try:
                    ref, created_ref = NormativeReference.objects.get_or_create(
                        source_standard=source,
                        cited_standard_no=cited_no,
                        defaults={
                            'cited_standard': cited,
                            'latest_standard_no': cited_no
                        }
                    )
                    if created_ref:
                        result['citations_added'] += 1
                        if cited:
                            citation_count_delta[cited.standard_no] = citation_count_delta.get(cited.standard_no, 0) + 1
                except Exception as e:
                    result['errors'].append(f'引用记录创建失败 {source_no}→{cited_no}: {str(e)}')

            for cited_no, delta in citation_count_delta.items():
                Standard.objects.filter(standard_no=cited_no, type='national').update(
                    citation_count=F('citation_count') + delta
                )

        # 5. 后置自动扫盘匹配一次 PDF 相对路径
        try:
            scan_res = scan_and_align_pdf_assets()
            result['pdf_aligned_count'] = scan_res.get('matched_count', 0)
        except Exception as scan_err:
            result['errors'].append(f"关联扫盘对齐 PDF 失败: {str(scan_err)}")

        return result


def get_citation_ranking(limit: int = 20):
    """
    获取国标引用热度排行榜（基于被引用的“发布时引用的完整标准号”进行累加统计，并关联展示“最新标准号”）
    """
    from django.db.models import Count, Max
    from standards.models import NormativeReference

    ranking = (
        NormativeReference.objects
        .values('cited_standard_no')
        .annotate(
            citation_count=Count('id'),
            latest_no=Max('latest_standard_no')
        )
        .order_by('-citation_count', 'cited_standard_no')[:limit]
    )

    results = []
    for idx, item in enumerate(ranking):
        cited_no = item['cited_standard_no']
        latest_no = item['latest_no'] or cited_no
        results.append({
            'id': idx + 1,
            'standard_no': cited_no,      # 发布时引用的完整标准号
            'title': latest_no,           # 最新标准号（替代原“标准名称”列显示）
            'citation_count': item['citation_count']
        })
    return results


# ============================================================
# 共享磁盘文件对齐与扫盘服务
# ============================================================

def normalize_string(s: str) -> str:
    """
    去除所有非字母数字字符并转为小写，用于模糊匹配标准号与文件名
    """
    if not s:
        return ""
    if s.lower().endswith('.pdf'):
        s = s[:-4]
    return re.sub(r'[^a-zA-Z0-9]', '', s).lower()


def scan_and_align_pdf_assets() -> dict:
    """
    自动遍历 Y:\磁盘阵列\标准文件下载\企标下载\整合\ 目录下的平铺 PDF 文件，
    对库内缺失 pdf_file 相对路径的企标进行多维子串模糊匹配对齐。
    """
    import os
    from django.conf import settings
    
    shared_root = getattr(settings, 'SHARED_DISK_ROOT', r"Y:\磁盘阵列\标准文件下载\企标下载")
    target_dir = os.path.join(shared_root, "整合")
    
    if not os.path.exists(target_dir):
        return {"success": False, "matched_count": 0, "error": f"物理磁盘路径不存在: {target_dir}"}
        
    # 1. 扫描整合目录下所有的 pdf 文件并建立归一化映射
    try:
        disk_files = [f for f in os.listdir(target_dir) if f.lower().endswith('.pdf')]
    except Exception as e:
        return {"success": False, "matched_count": 0, "error": f"读取磁盘目录失败: {str(e)}"}
        
    normalized_file_map = {normalize_string(f): f for f in disk_files}
    
    # 2. 查询缺失文件或物理路径失效的企标进行对齐
    all_standards = Standard.objects.filter(type='enterprise')
    unlinked_standards = []
    
    for std in all_standards:
        if not std.pdf_file or not std.pdf_file.name:
            unlinked_standards.append(std)
            continue
            
        # 检查物理文件是否真实存在于本地或共享盘中
        rel_path = std.pdf_file.name
        
        # 本地 media 路径
        local_exists = os.path.exists(os.path.join(settings.MEDIA_ROOT, rel_path))
        # 共享磁盘路径
        shared_exists = os.path.exists(os.path.join(shared_root, rel_path))
        
        # 兜底清理 media/ 前缀检查
        if not shared_exists and rel_path.startswith('media/'):
            clean_path = rel_path.replace('media/', '', 1)
            local_exists = local_exists or os.path.exists(os.path.join(settings.MEDIA_ROOT, clean_path))
            
        if not local_exists and not shared_exists:
            # 物理文件缺失，需要扫盘重新匹配对齐
            unlinked_standards.append(std)
    
    matched_count = 0
    for std in unlinked_standards:
        std_norm = normalize_string(std.standard_no)
        matched_filename = None
        
        # 2a. 精确匹配
        if std_norm in normalized_file_map:
            matched_filename = normalized_file_map[std_norm]
        else:
            # 2b. 模糊包含匹配
            for file_norm, filename in normalized_file_map.items():
                if std_norm in file_norm or file_norm in std_norm:
                    matched_filename = filename
                    break
                    
        if matched_filename:
            std.pdf_file = f"整合/{matched_filename}"
            std.save(update_fields=['pdf_file'])
            matched_count += 1
            
    return {"success": True, "matched_count": matched_count}


# ============================================================
# ZIP 打包服务（Celery 异步任务辅助）
# ============================================================

from standards.utils.archive_helpers import create_zip_from_standards


# ============================================================
# Excel 批量导入企标（按照爬取 Excel 表头与核心业务逻辑）
# ============================================================

def import_standards_from_excel(file_obj) -> dict:
    """
    从爬取的 Excel 文件批量导入企标和企业信息（高度融合核心业务与 LBS 定位）

    适合 crawled Excel 表头格式，并自动对齐去重、物理路径分层与直辖市二级结构兼容。
    """
    import pandas as pd
    from django.utils import timezone
    from companies.models import Company, Province, City, District
    
    result = {
        'success': 0,
        'skipped': 0,
        'companies_created': 0,
        'companies_updated': 0,
        'errors': []
    }
    
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        result['errors'].append(f'文件解析失败: {str(e)}')
        return result
        
    today_str = timezone.now().strftime("%Y/%m/%d")
    
    standards_to_create = []
    companies_created_count = 0
    companies_updated_count = 0
    skipped_duplicates_count = 0
    
    with transaction.atomic():
        for idx, row in df.iterrows():
            credit_code = str(row.get('统一社会信用代码', '')).strip()
            company_name = str(row.get('起草单位/企业名称', '')).strip()
            
            if not credit_code or credit_code == 'nan' or not company_name or company_name == 'nan':
                continue
                
            # --- 步骤 A: 级联匹配省市区外键 ---
            province = None
            city = None
            district = None
            
            geo_str = str(row.get('行政区划', '')).strip()
            if geo_str and geo_str != 'nan' and '-' in geo_str:
                geo_parts = geo_str.split('-')
                if len(geo_parts) >= 1:
                    province = Province.objects.filter(name__icontains=geo_parts[0].replace("省", "")).first()
                
                # 处理直辖市两级行政区划 (如北京市-海淀区)
                if len(geo_parts) == 2 and province:
                    city = City.objects.filter(name__in=["市辖区", province.name.replace("市", "")], province=province).first()
                    if not city:
                        city = City.objects.filter(province=province).first()
                    if city:
                        district = District.objects.filter(name__icontains=geo_parts[1], city=city).first()
                else:
                    # 标准三级区划结构 (如安徽省-安庆市-桐城市)
                    if len(geo_parts) >= 2 and province:
                        city = City.objects.filter(name__icontains=geo_parts[1].replace("市", ""), province=province).first()
                    if len(geo_parts) >= 3 and city:
                        district = District.objects.filter(name__icontains=geo_parts[2], city=city).first()
            
            # --- 步骤 B: 根据大字典自动获取 LBS 中心经纬度 ---
            company_lat = None
            company_lng = None
            if district and district.latitude and district.longitude:
                company_lat = district.latitude
                company_lng = district.longitude
            elif city and city.latitude and city.longitude:
                company_lat = city.latitude
                company_lng = city.longitude

            # --- 步骤 C: 新建或安全更新企业信息（按照信用代码去重） ---
            legal_person = str(row.get('法定代表人', '')).strip()
            if not legal_person or legal_person == 'nan':
                legal_person = ''
                
            address = str(row.get('注册地址', '')).strip()
            if not address or address == 'nan':
                address = ''

            company, created = Company.objects.get_or_create(
                credit_code=credit_code,
                defaults={
                    'name': company_name,
                    'legal_person': legal_person,
                    'address': address,
                    'province': province,
                    'city': city,
                    'district': district,
                    'latitude': company_lat,
                    'longitude': company_lng,
                    'status': 'active'
                }
            )
            
            if created:
                companies_created_count += 1
            else:
                updated = False
                if not company.latitude and company_lat:
                    company.latitude = company_lat
                    company.longitude = company_lng
                    updated = True
                if not company.district and district:
                    company.province = province
                    company.city = city
                    company.district = district
                    updated = True
                if updated:
                    company.save()
                    companies_updated_count += 1
            
            # --- 步骤 D: 拼装标准记录并建立外键绑定 ---
            standard_no = str(row.get('标准编号', '')).strip()
            standard_title = str(row.get('标准名称', '')).strip()
            if standard_title.startswith('《') and standard_title.endswith('》'):
                standard_title = standard_title[1:-1].strip()
            
            if not standard_no or standard_no == 'nan':
                continue
                
            clean_id = generate_clean_id(standard_no)
            
            if Standard.objects.filter(clean_id=clean_id).exists():
                skipped_duplicates_count += 1
                continue
                
            publish_date = None
            pub_date_val = row.get('公开时间')
            if pd.notnull(pub_date_val):
                try:
                    publish_date = pd.to_datetime(pub_date_val).date()
                except:
                    pass
            
            status_str = str(row.get('标准状态', ''))
            standard_status = 'active'
            if '废止' in status_str:
                standard_status = 'deprecated'
            elif '草案' in status_str:
                standard_status = 'draft'

            pdf_filename = str(row.get('PDF文件名', '')).strip()
            pdf_db_path = f"整合/{pdf_filename}" if pdf_filename and pdf_filename != 'nan' else ''
            
            new_standard = Standard(
                standard_no=standard_no,
                clean_id=clean_id,
                type='enterprise',
                title=standard_title if standard_title != 'nan' else '',
                company=company,
                pdf_file=pdf_db_path,
                publish_date=publish_date,
                status=standard_status
            )
            standards_to_create.append(new_standard)
            
        if standards_to_create:
            Standard.objects.bulk_create(standards_to_create, batch_size=1000)
            from django.core.cache import cache
            from standards.models import _invalidate_search_cache
            try:
                _invalidate_search_cache()
            except Exception:
                pass
            
    # 自动进行 Y:\ 磁盘扫盘匹配对齐，补全 relative_path (pdf_file)
    try:
        scan_res = scan_and_align_pdf_assets()
        result['pdf_aligned_count'] = scan_res.get('matched_count', 0)
    except Exception:
        pass
            
    result['success'] = len(standards_to_create)
    result['skipped'] = skipped_duplicates_count
    result['companies_created'] = companies_created_count
    result['companies_updated'] = companies_updated_count
    return result


def generate_standard_import_template(std_type: str) -> tuple:
    """
    动态生成导入 Excel 模板，返回 (excel_bytes, filename)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '标准导入模板'
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if std_type == 'enterprise':
        headers = ['标准编号*', '标准名称*', '起草单位/企业名称*', '统一社会信用代码*', '法定代表人', '注册地址', '公开时间(YYYY-MM-DD)', '标准状态(现行/废止/草案)', 'PDF文件名']
        filename = "企业标准导入模板.xlsx"
    else:
        # 国标/行标/地标/团标模板
        headers = ['标准编号*', '标准名称*', '发布日期(YYYY-MM-DD)', '实施日期(YYYY-MM-DD)', '标准状态(现行/废止/草案)', 'ICS分类号', 'CCS分类号', '关联起草企业名称(可选)', '起草企业信用代码(可选)', '磁盘阵列文件路径(可选)']
        type_map = {
            'national': '国家标准',
            'industry': '行业标准',
            'local': '地方标准',
            'group': '团体标准'
        }
        filename = f"{type_map.get(std_type, '其他标准')}导入模板.xlsx"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename


def import_standards_by_type(file_obj, std_type: str) -> dict:
    """
    多类型标准统一导入逻辑。
    如果是 enterprise（企业标准），保留原有逻辑，实现对起草单位的企业建档以及 PDF 扫盘对齐。
    如果是 national/industry/local/group 等，支持字段解析，并支持可选的企业关联。
    """
    if std_type == 'enterprise':
        return import_standards_from_excel(file_obj)

    # 非企业标准导入逻辑
    import pandas as pd
    import uuid
    from django.utils import timezone
    from companies.models import Company
    from standards.models import Standard
    
    result = {
        'success': 0,
        'skipped': 0,
        'companies_created': 0,
        'companies_updated': 0,
        'errors': []
    }

    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        result['errors'].append(f'文件解析失败: {str(e)}')
        return result

    if df.empty:
        result['errors'].append('Excel 数据为空')
        return result

    standards_to_create = []
    skipped_count = 0

    with transaction.atomic():
        for idx, row in df.iterrows():
            row_idx = idx + 2
            
            # 安全读取前两个核心必填字段
            std_no = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            std_title = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            
            if not std_no or std_no == 'nan':
                result['errors'].append(f'第 {row_idx} 行: 标准编号必填，已跳过')
                continue
            if not std_title or std_title == 'nan':
                result['errors'].append(f'第 {row_idx} 行: 标准名称必填，已跳过')
                continue

            clean_id = generate_clean_id(std_no)

            # 查重
            if Standard.objects.filter(clean_id=clean_id).exists():
                skipped_count += 1
                continue

            # 日期解析
            publish_date = None
            implement_date = None
            if len(row) > 2 and pd.notnull(row.iloc[2]):
                try: publish_date = pd.to_datetime(row.iloc[2]).date()
                except: pass
            if len(row) > 3 and pd.notnull(row.iloc[3]):
                try: implement_date = pd.to_datetime(row.iloc[3]).date()
                except: pass

            # 状态解析
            status_str = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ''
            standard_status = 'active'
            if '废止' in status_str:
                standard_status = 'deprecated'
            elif '草案' in status_str:
                standard_status = 'draft'

            ics_val = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else ''
            ccs_val = str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else ''

            # 关联企业逻辑
            company = None
            co_name = str(row.iloc[7]).strip() if len(row) > 7 and pd.notna(row.iloc[7]) else ''
            co_code = str(row.iloc[8]).strip() if len(row) > 8 and pd.notna(row.iloc[8]) else ''

            if co_code and co_code != 'nan':
                company = Company.objects.filter(credit_code=co_code).first()
                if not company and co_name and co_name != 'nan':
                    company = Company.objects.create(
                        name=co_name,
                        credit_code=co_code,
                        status='active'
                    )
                    result['companies_created'] += 1
            elif co_name and co_name != 'nan':
                company = Company.objects.filter(name=co_name).first()
                if not company:
                    company = Company.objects.create(
                        name=co_name,
                        credit_code=f'TEMP_{uuid.uuid4().hex[:14].upper()}',
                        status='active'
                    )
                    result['companies_created'] += 1

            disk_path = str(row.iloc[9]).strip() if len(row) > 9 and pd.notna(row.iloc[9]) else ''

            new_standard = Standard(
                standard_no=std_no,
                clean_id=clean_id,
                type=std_type,
                title=std_title,
                company=company,
                publish_date=publish_date,
                implement_date=implement_date,
                status=standard_status,
                ics=ics_val if ics_val != 'nan' else '',
                ccs=ccs_val if ccs_val != 'nan' else '',
                disk_filename=disk_path if disk_path != 'nan' else ''
            )
            standards_to_create.append(new_standard)

        if standards_to_create:
            Standard.objects.bulk_create(standards_to_create, batch_size=1000)
            from standards.models import _invalidate_search_cache
            try:
                _invalidate_search_cache()
            except Exception:
                pass

    result['success'] = len(standards_to_create)
    result['skipped'] = skipped_count
    return result


def generate_reference_import_template_v2() -> tuple:
    """
    动态生成规范性引用导入 Excel 模板，返回 (excel_bytes, filename)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '规范性引用导入模板'
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['企标编号*', '引用的国标/行标编号*', '被引用标准名称', '最新标准号']
    filename = "企标规范性引用导入模板.xlsx"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename


def import_references_from_excel_v2(file_obj) -> dict:
    """
    后台导入企标规范性引用，包含严格的行级数据校验
    对正确的数据执行入库，并收集所有错误反馈给前端
    """
    import pandas as pd
    from django.db import transaction
    from django.db.models import F
    from standards.models import Standard, NormativeReference
    
    result = {
        'success_count': 0,
        'errors': []
    }
    
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        result['errors'].append({'row': 0, 'error': f'文件解析失败: {str(e)}'})
        return result
        
    if df.empty:
        result['errors'].append({'row': 0, 'error': 'Excel 文件内容为空'})
        return result
        
    # 必要表头校验
    expected_headers = ['企标编号*', '引用的国标/行标编号*']
    for header in expected_headers:
        if header not in df.columns:
            clean_header = header.replace('*', '')
            found = False
            for col in df.columns:
                if clean_header in str(col):
                    df.rename(columns={col: header}, inplace=True)
                    found = True
                    break
            if not found:
                result['errors'].append({'row': 1, 'error': f"Excel 模板格式错误，缺失必要列: '{header}'"})
                return result

    # 识别并重命名“最新标准号”列（防止用户上传的 Excel 包含空格、星号等微小字样差异）
    latest_header = '最新标准号'
    if latest_header not in df.columns:
        for col in df.columns:
            if '最新标准号' in str(col) or '最新被引用标准号' in str(col):
                df.rename(columns={col: latest_header}, inplace=True)
                break

    # 1. 缓存加载全部输入企标，防止 N+1 查询挂起
    source_nos = df['企标编号*'].dropna().astype(str).str.strip().unique().tolist()
    standard_map = {
        std.standard_no: std
        for std in Standard.objects.filter(standard_no__in=source_nos, type='enterprise')
    }
    
    # 2. 缓存已存在的国/行标以便匹配
    cited_nos = df['引用的国标/行标编号*'].dropna().astype(str).str.strip().unique().tolist()
    cited_std_map = {
        std.standard_no: std
        for std in Standard.objects.filter(standard_no__in=cited_nos)
    }

    # 3. 逐行处理及校验
    with transaction.atomic():
        for idx, row in df.iterrows():
            row_idx = idx + 2  # Excel 实际数据从第 2 行开始
            
            source_no = str(row.get('企标编号*', '')).strip() if pd.notna(row.get('企标编号*', '')) else ''
            cited_no = str(row.get('引用的国标/行标编号*', '')).strip() if pd.notna(row.get('引用的国标/行标编号*', '')) else ''
            latest_no = str(row.get('最新标准号', '')).strip() if pd.notna(row.get('最新标准号', '')) else ''
            
            # A. 必填字段格式校验
            if not source_no or source_no == 'nan':
                result['errors'].append({'row': row_idx, 'error': "必填列 '企标编号*' 为空"})
                continue
            if not cited_no or cited_no == 'nan':
                result['errors'].append({'row': row_idx, 'error': "必填列 '引用的国标/行标编号*' 为空"})
                continue
                
            # B. 主体存在性校验
            standard = standard_map.get(source_no)
            if not standard:
                result['errors'].append({
                    'row': row_idx, 
                    'error': f"企标编号 '{source_no}' 在系统企标库中未找到，请确保已一键导入该企标主表"
                })
                continue
                
            # C. 关联引用的国行标
            cited_std = cited_std_map.get(cited_no)
            
            try:
                ref, created = NormativeReference.objects.get_or_create(
                    source_standard=standard,
                    cited_standard_no=cited_no,
                    defaults={
                        'cited_standard': cited_std,
                        'latest_standard_no': latest_no or cited_no
                    }
                )
                if created:
                    result['success_count'] += 1
                    # 联动更新企标的解析状态为 references_parsed
                    if standard.is_parsed == 'unparsed':
                        standard.is_parsed = 'references_parsed'
                        standard.save(update_fields=['is_parsed'])
                    # 联动增加国行标被引用计数
                    if cited_std:
                        Standard.objects.filter(pk=cited_std.pk).update(citation_count=F('citation_count') + 1)
                else:
                    if latest_no and ref.latest_standard_no != latest_no:
                        ref.latest_standard_no = latest_no
                        ref.save(update_fields=['latest_standard_no'])
            except Exception as e:
                result['errors'].append({'row': row_idx, 'error': f"数据保存入库失败: {str(e)}"})

    # 后置扫盘匹配对齐 PDF
    try:
        scan_and_align_pdf_assets()
    except Exception:
        pass
        
    return result


def generate_mixed_import_template_v2() -> tuple:
    """
    动态生成企标与规范引用混合导入模板 Excel，返回 (excel_bytes, filename)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '企标与引用混合导入模板'
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['企标编号*', '企标名称*', '起草单位*', '统一社会信用代码*', '引用的国标/行标编号*', '最新标准号']
    filename = "企业标准与引用关系混合导入模板.xlsx"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename
