import os
import traceback
import re
import datetime
from celery import shared_task
from django.core.cache import cache
from django.utils import timezone
import openpyxl
from django.conf import settings

from .models import Company, Province, City, District


def _update_progress(task_id, status, progress, success=0, skipped=0, errors=None, total=0):
    if errors is None:
        errors = []
    cache_key = f"import_task_{task_id}"
    cache.set(cache_key, {
        "status": status,
        "progress": progress,
        "success": success,
        "skipped": skipped,
        "errors": errors[:100],  # 只保留前 100 个错误避免缓存过大
        "total": total,
    }, timeout=3600)


@shared_task
def import_companies_task(file_path, task_id):
    """
    异步处理 Excel 批量导入企业任务
    1. 读预热缓存（省市区 3 次 DB → 0 次）
    2. 第一遍流式读取 Excel 提取所有信用代码，一次性加载库中已存记录，防 N+1
    3. 支持覆盖更新已存在的企业，补全所有 20+ 个详细字段
    4. 分别执行 bulk_create 和 bulk_update 进行批量高速入库
    """
    _update_progress(task_id, status="processing", progress=1)
    errors = []
    success_count = 0
    skipped_count = 0
    total_count = 0

    try:
        # ==========================================
        # Step 1: 从预热缓存加载省市区字典
        # ==========================================
        _update_progress(task_id, status="processing", progress=5)

        from companies.warmup import get_area_data_from_cache
        _provinces, _cities, _districts = get_area_data_from_cache()

        # 将列表转为查找字典
        prov_dict = {p['name']: p for p in _provinces}
        city_dict  = {(c['province_id'], c['name']): c for c in _cities}
        dist_dict  = {(d['city_id'], d['name']): d for d in _districts}

        def find_province(name):
            if not name: return None
            for p_name, p_data in prov_dict.items():
                if name in p_name or p_name in name:
                    return p_data
            return None

        def find_city(p_id, name):
            if not name or not p_id: return None
            for (pid, c_name), c_data in city_dict.items():
                if pid == p_id and (name in c_name or c_name in name):
                    return c_data
            return None

        def find_district(c_id, name):
            if not name or not c_id: return None
            for (cid, d_name), d_data in dist_dict.items():
                if cid == c_id and (name in d_name or d_name in name):
                    return d_data
            return None

        # ==========================================
        # Step 2: 第一遍读取 Excel，快速收集信用代码以进行批量对齐
        # ==========================================
        _update_progress(task_id, status="processing", progress=8)
        excel_credit_codes = set()
        wb_temp = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws_temp = wb_temp.active
        rows_temp = ws_temp.iter_rows(values_only=True)
        try:
            h_row = next(rows_temp)
        except StopIteration:
            h_row = []

        cc_idx = -1
        if h_row:
            headers_temp = [str(c).strip() if c else '' for c in h_row]
            for i, h in enumerate(headers_temp):
                if h in ['统一社会信用代码*', '统一社会信用代码']:
                    cc_idx = i
                    break
            if cc_idx == -1:
                cc_idx = 1  # 降级匹配第二列

            for r in rows_temp:
                if r and len(r) > cc_idx:
                    val = str(r[cc_idx]).strip() if r[cc_idx] else ''
                    if val and val != 'nan':
                        excel_credit_codes.add(val)
        wb_temp.close()

        # 批量加载已存在且未删除的企业档案
        existing_companies = {}
        if excel_credit_codes:
            existing_companies = {
                c.credit_code: c for c in Company.objects.filter(
                    credit_code__in=excel_credit_codes,
                    is_deleted=False
                )
            }

        # ==========================================
        # Step 3: 二次读取，逐行进行深度解析与更新
        # ==========================================
        _update_progress(task_id, status="processing", progress=12)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            _update_progress(task_id, status="failed", progress=100, errors=["文件为空"])
            return

        headers = [str(c).strip() if c else '' for c in header_row]
        col_map = {
            'name': -1, 'credit_code': -1, 'legal_person': -1,
            'province': -1, 'city': -1, 'district': -1,
            'region_full': -1, 'latitude': -1, 'longitude': -1,
            'contact': -1, 'address': -1,
            # 扩展字段映射
            'established_date': -1, 'registered_address': -1, 'registered_zipcode': -1,
            'valid_mobile': -1, 'more_phones': -1, 'email': -1,
            'company_type': -1, 'registration_no': -1, 'organization_code': -1,
            'industry_category': -1, 'industry_major': -1, 'industry_middle': -1, 'industry_minor': -1,
            'company_size': -1, 'english_name': -1, 'former_names': -1,
            'website_url': -1, 'mailing_address': -1, 'mailing_address_zip': -1,
            'business_scope': -1, 'registration_status': -1
        }

        for i, h in enumerate(headers):
            if h in ['企业名称*', '企业名称', '起草单位/企业名称']:       col_map['name'] = i
            elif h in ['统一社会信用代码*', '统一社会信用代码']:           col_map['credit_code'] = i
            elif h in ['法人', '法定代表人']:                             col_map['legal_person'] = i
            elif h in ['省份', '所属省份']:                              col_map['province'] = i
            elif h in ['城市', '所属城市']:                              col_map['city'] = i
            elif h in ['区县', '所属区县']:                              col_map['district'] = i
            elif h in ['行政区划']:                                       col_map['region_full'] = i
            elif h in ['纬度', '实际经纬度']:                             col_map['latitude'] = i
            elif h in ['经度']:                                          col_map['longitude'] = i
            elif h in ['联系方式', '负责人电话/手机号码']:                 col_map['contact'] = i
            elif h in ['详细地址', '注册地址', '实际地址']:                col_map['address'] = i
            # 扩展与补充字段
            elif h in ['成立日期', '成立时间']:                           col_map['established_date'] = i
            elif h in ['营业执照住所', '注册住所地址', '注册地址']:         col_map['registered_address'] = i
            elif h in ['注册地址邮编', '注册邮编']:                        col_map['registered_zipcode'] = i
            elif h in ['有效手机号', '主要联系手机']:                      col_map['valid_mobile'] = i
            elif h in ['更多电话', '其他备用电话']:                        col_map['more_phones'] = i
            elif h in ['邮箱', '电子邮箱', '企业邮箱']:                    col_map['email'] = i
            elif h in ['企业(机构)类型', '企业类型', '机构类型']:          col_map['company_type'] = i
            elif h in ['注册号', '工商注册号']:                           col_map['registration_no'] = i
            elif h in ['组织机构代码']:                                   col_map['organization_code'] = i
            elif h in ['国标行业门类', '行业门类']:                       col_map['industry_category'] = i
            elif h in ['国标行业大类', '行业大类']:                       col_map['industry_major'] = i
            elif h in ['国标行业中类', '行业中类']:                       col_map['industry_middle'] = i
            elif h in ['国标行业小类', '行业小类']:                       col_map['industry_minor'] = i
            elif h in ['企业规模']:                                       col_map['company_size'] = i
            elif h in ['英文名', '企业英文名称']:                         col_map['english_name'] = i
            elif h in ['曾用名', '历史曾用名']:                           col_map['former_names'] = i
            elif h in ['官网网址', '官网地址', '网址']:                     col_map['website_url'] = i
            elif h in ['通讯地址']:                                       col_map['mailing_address'] = i
            elif h in ['通讯地址邮编', '落地地址邮编']:                    col_map['mailing_address_zip'] = i
            elif h in ['经营范围', '工商经营范围']:                       col_map['business_scope'] = i
            elif h in ['工商状态', '登记状态']:                           col_map['registration_status'] = i

        # 列头全未识别时按位置降级匹配
        if all(v == -1 for v in col_map.values() if v not in [-1]):
            col_map = {
                'name': 0, 'credit_code': 1, 'legal_person': 2,
                'province': 3, 'city': 4, 'district': 5,
                'latitude': 6, 'longitude': 7, 'contact': 8,
                'address': 9, 'region_full': -1
            }

        estimated_total = ws.max_row if ws.max_row else 10000

        # ==========================================
        # Step 4: 循环读取每行并分类存入 create/update 队列
        # ==========================================
        batch_size = 2000
        companies_to_create = []
        companies_to_update = []

        for row_idx, row in enumerate(rows_iter, start=2):
            if not any(row):
                continue
            total_count += 1

            def get_val(key):
                idx = col_map.get(key, -1)
                if idx != -1 and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ''
                return ''

            name        = get_val('name')
            credit_code = get_val('credit_code')

            if not name:
                errors.append(f"第{row_idx}行: 企业名称不能为空")
                continue
            if not credit_code:
                errors.append(f"第{row_idx}行: 统一社会信用代码不能为空")
                continue

            # 行政区划解析
            p_data = c_data = d_data = None
            region_full = get_val('region_full')

            if region_full and region_full != 'nan' and '-' in region_full:
                geo_parts = region_full.split('-')
                if len(geo_parts) >= 1:
                    p_data = find_province(geo_parts[0].replace('省', ''))
                if len(geo_parts) == 2 and p_data:
                    c_data = (find_city(p_data['id'], '市辖区')
                              or find_city(p_data['id'], p_data['name'].replace('省', '').replace('市', '')))
                    if c_data:
                        d_data = find_district(c_data['id'], geo_parts[1])
                else:
                    if len(geo_parts) >= 2 and p_data:
                        c_data = find_city(p_data['id'], geo_parts[1].replace('市', ''))
                    if len(geo_parts) >= 3 and c_data:
                        d_data = find_district(c_data['id'], geo_parts[2])

            if not p_data:
                p_data = find_province(get_val('province'))
            if not c_data and p_data:
                c_data = find_city(p_data['id'], get_val('city'))
            if not d_data and c_data:
                d_data = find_district(c_data['id'], get_val('district'))

            # 经纬度解析（支持经度,纬度合并字段或单列字段）
            latitude = longitude = None
            try:
                lat_str = get_val('latitude')
                lng_str = get_val('longitude')
                if lat_str and (',' in lat_str or ';' in lat_str or '，' in lat_str):
                    parts = re.split(r'[,;，]+', lat_str)
                    if len(parts) >= 2:
                        val1 = float(parts[0].strip())
                        val2 = float(parts[1].strip())
                        # 根据典型经纬度范围自动纠正经纬度顺序
                        if 70 <= val1 <= 140 and 3 <= val2 <= 60:
                            longitude = val1
                            latitude = val2
                        elif 70 <= val2 <= 140 and 3 <= val1 <= 60:
                            longitude = val2
                            latitude = val1
                        else:
                            latitude = val1
                            longitude = val2
                else:
                    if lat_str: latitude = float(lat_str)
                    if lng_str: longitude = float(lng_str)
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行: 经纬度格式有误")

            if latitude is None and longitude is None:
                if d_data and d_data.get('latitude') and d_data.get('longitude'):
                    latitude  = d_data['latitude']
                    longitude = d_data['longitude']
                elif c_data and c_data.get('latitude') and c_data.get('longitude'):
                    latitude  = c_data['latitude']
                    longitude = c_data['longitude']

            # 成立日期时间解析与安全校验
            est_date_str = get_val('established_date')
            established_date = None
            if est_date_str and est_date_str != 'nan':
                try:
                    cleaned_date = est_date_str.replace('/', '-').split(' ')[0]
                    established_date = datetime.datetime.strptime(cleaned_date, '%Y-%m-%d').date()
                except Exception:
                    try:
                        established_date = datetime.datetime.strptime(est_date_str[:10], '%Y-%m-%d').date()
                    except Exception:
                        errors.append(f"第{row_idx}行: 成立日期格式有误 ('{est_date_str}')，应为 YYYY-MM-DD")

            # 提取详细信息以合并到公司模型中（仅在非空时覆盖）
            def assign_fields(c):
                c.name = name
                if get_val('legal_person'): c.legal_person = get_val('legal_person')
                if p_data: c.province_id = p_data['id']
                if c_data: c.city_id = c_data['id']
                if d_data: c.district_id = d_data['id']
                if latitude is not None: c.latitude = latitude
                if longitude is not None: c.longitude = longitude
                if get_val('contact'): c.contact = get_val('contact')
                if get_val('address'): c.address = get_val('address')
                
                # 扩展与补充字段赋值
                if established_date: c.established_date = established_date
                if get_val('registered_address'): c.registered_address = get_val('registered_address')
                if get_val('registered_zipcode'): c.registered_zipcode = get_val('registered_zipcode')
                if get_val('valid_mobile'): c.valid_mobile = get_val('valid_mobile')
                if get_val('more_phones'): c.more_phones = get_val('more_phones')
                if get_val('email'): c.email = get_val('email')
                if get_val('company_type'): c.company_type = get_val('company_type')
                if get_val('registration_no'): c.registration_no = get_val('registration_no')
                if get_val('organization_code'): c.organization_code = get_val('organization_code')
                if get_val('industry_category'): c.industry_category = get_val('industry_category')
                if get_val('industry_major'): c.industry_major = get_val('industry_major')
                if get_val('industry_middle'): c.industry_middle = get_val('industry_middle')
                if get_val('industry_minor'): c.industry_minor = get_val('industry_minor')
                if get_val('company_size'): c.company_size = get_val('company_size')
                if get_val('english_name'): c.english_name = get_val('english_name')
                if get_val('former_names'): c.former_names = get_val('former_names')
                if get_val('website_url'): c.website_url = get_val('website_url')
                if get_val('mailing_address'): c.mailing_address = get_val('mailing_address')
                if get_val('mailing_address_zip'): c.mailing_address_zip = get_val('mailing_address_zip')
                if get_val('business_scope'): c.business_scope = get_val('business_scope')
                if get_val('registration_status'): c.registration_status = get_val('registration_status')
                
                c.updated_at = timezone.now()

            # 判断是覆盖更新还是新建
            if credit_code in existing_companies:
                company = existing_companies[credit_code]
                assign_fields(company)
                companies_to_update.append(company)
            else:
                company = Company(
                    credit_code=credit_code,
                    status='active',
                    is_deleted=False,
                    created_at=timezone.now()
                )
                assign_fields(company)
                companies_to_create.append(company)
                existing_companies[credit_code] = company

            # 批量写入与更新
            if len(companies_to_create) >= batch_size:
                Company.objects.bulk_create(companies_to_create, ignore_conflicts=True)
                success_count += len(companies_to_create)
                companies_to_create = []

            if len(companies_to_update) >= batch_size:
                Company.objects.bulk_update(companies_to_update, fields=[
                    'name', 'legal_person', 'province_id', 'city_id', 'district_id',
                    'latitude', 'longitude', 'contact', 'address', 'established_date',
                    'registered_address', 'registered_zipcode', 'valid_mobile',
                    'more_phones', 'email', 'company_type', 'registration_no',
                    'organization_code', 'industry_category', 'industry_major',
                    'industry_middle', 'industry_minor', 'company_size', 'english_name',
                    'former_names', 'website_url', 'mailing_address', 'mailing_address_zip',
                    'business_scope', 'registration_status', 'updated_at'
                ])
                success_count += len(companies_to_update)
                companies_to_update = []

            if total_count % 100 == 0:
                progress = min(12 + int((total_count / estimated_total) * 83), 95)
                _update_progress(task_id, status="processing", progress=progress,
                                 success=success_count, skipped=skipped_count,
                                 errors=errors, total=total_count)

        # 插入与更新剩余数据
        if companies_to_create:
            Company.objects.bulk_create(companies_to_create, ignore_conflicts=True)
            success_count += len(companies_to_create)

        if companies_to_update:
            Company.objects.bulk_update(companies_to_update, fields=[
                'name', 'legal_person', 'province_id', 'city_id', 'district_id',
                'latitude', 'longitude', 'contact', 'address', 'established_date',
                'registered_address', 'registered_zipcode', 'valid_mobile',
                'more_phones', 'email', 'company_type', 'registration_no',
                'organization_code', 'industry_category', 'industry_major',
                'industry_middle', 'industry_minor', 'company_size', 'english_name',
                'former_names', 'website_url', 'mailing_address', 'mailing_address_zip',
                'business_scope', 'registration_status', 'updated_at'
            ])
            success_count += len(companies_to_update)

        wb.close()

        # 删除临时上传文件
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

        _update_progress(task_id, status="done", progress=100,
                          success=success_count, skipped=skipped_count,
                          errors=errors, total=total_count)

    except Exception as e:
        traceback.print_exc()
        errors.append(f"系统错误: {str(e)}")
        _update_progress(task_id, status="failed", progress=100,
                         success=success_count, skipped=skipped_count,
                         errors=errors, total=total_count)


# ============================================================
# Celery Beat 定时预热任务
# ============================================================

@shared_task(name='companies.tasks.warm_area_dict_task')
def warm_area_dict_task():
    """
    Celery Beat 定时任务：每 24 小时刷新省市区字典缓存。
    对应 settings.CELERY_BEAT_SCHEDULE 中的 'warm-area-dict-daily'。
    """
    from companies.warmup import warm_area_dict
    warm_area_dict()
    return 'area dict cache warmed'


@shared_task(name='companies.tasks.warm_dashboard_stats_task')
def warm_dashboard_stats_task():
    """
    Celery Beat 定时任务：每 30 分钟刷新 Dashboard 统计数据缓存。
    对应 settings.CELERY_BEAT_SCHEDULE 中的 'warm-dashboard-stats-30m'。
    """
    try:
        total_companies  = Company.objects.filter(is_deleted=False).count()
        active_companies = Company.objects.filter(is_deleted=False, status='active').count()
        cache.set('dashboard:stats', {
            'total_companies':  total_companies,
            'active_companies': active_companies,
        }, timeout=1800)  # 30 分钟
    except Exception as e:
        import logging
        logging.getLogger('companies.tasks').warning(f'[warm_dashboard] 失败: {e}')
    return 'dashboard stats cache warmed'
