"""
companies.services — 企业业务逻辑层

按"胖 Model / 瘦 View"原则，复杂逻辑不写在 views.py 中。

功能：
  - 多维检索（行政/LBS/关键词）
  - Excel 批量导入（含查重校验）
  - Excel 导出
"""

import math
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction, IntegrityError
from django.db.models import Q

from .models import Company, Province, City, District


# ============================================================
# LBS 空间检索（MySQL ST_Distance_Sphere）
# ============================================================

def search_companies_by_lbs(center_lat: float, center_lng: float, radius_km: float):
    """
    基于经纬度和半径（km）检索范围内的企业。

    使用 MySQL 8.x 内置 ST_Distance_Sphere 函数，
    不依赖 PostGIS，返回在圆形范围内的企业 QuerySet。

    Args:
        center_lat: 中心点纬度
        center_lng: 中心点经度
        radius_km:  搜索半径（公里）

    Returns:
        带距离注解的 Company QuerySet
    """
    from django.db.models.expressions import RawSQL
    from django.db.models import FloatField

    radius_meters = radius_km * 1000

    # MySQL 8.x: ST_Distance_Sphere(POINT(lng, lat), POINT(center_lng, center_lat))
    distance_expr = RawSQL(
        "ST_Distance_Sphere(POINT(%s, %s), POINT(longitude, latitude))",
        (center_lng, center_lat),
        output_field=FloatField()
    )

    return (
        Company.objects
        .filter(is_deleted=False, status='active')
        .exclude(latitude__isnull=True)
        .exclude(longitude__isnull=True)
        .annotate(distance_meters=distance_expr)
        .filter(distance_meters__lte=radius_meters)
        .order_by('distance_meters')
    )


# ============================================================
# 多维综合检索
# ============================================================

def search_companies(
    keyword: str = '',
    province_id: int = None,
    city_id: int = None,
    district_id: int = None,
    center_lat: float = None,
    center_lng: float = None,
    radius_km: float = None,
    ics: str = '',
    ccs: str = '',
    standard_logic: str = 'OR',
    status: str = 'active',
    category_id: int = None,
    category_code: str = '',
):
    """
    多维度企业与标准级联检索（支持任意组合条件）

    支持：
      1. 行政区划级联 (Province -> City -> District)
      2. 经纬度 LBS 范围筛选 (lat, lng, radius_km) 并能与其他条件 AND 联动
      3. 企业名称/信用代码关键词匹配
      4. 企业标准 ICS / CCS 分类筛选 (支持 OR / AND 逻辑)
      5. 企业所有制分类与标签树筛选 (大类/小类)
    """
    qs = Company.objects.filter(is_deleted=False)

    if status:
        qs = qs.filter(status=status)

    # 1. 行政区划筛选（三级级联 AND 逻辑）
    def to_int(val):
        try:
            return int(val) if val else None
        except (ValueError, TypeError):
            return None

    p_id = to_int(province_id)
    c_id = to_int(city_id)
    d_id = to_int(district_id)

    if p_id:
        qs = qs.filter(province_id=p_id)
    if c_id:
        qs = qs.filter(city_id=c_id)
    if d_id:
        qs = qs.filter(district_id=d_id)

    # 2. 所有制分类与标签筛选
    from .models import CompanyCategory
    cat_id = to_int(category_id)
    if cat_id:
        cat = CompanyCategory.objects.filter(id=cat_id).first()
        if cat:
            if cat.category_type == 'main':
                sub_ids = list(cat.children.values_list('id', flat=True))
                qs = qs.filter(ownership_categories__id__in=[cat.id] + sub_ids).distinct()
            else:
                qs = qs.filter(ownership_categories=cat).distinct()
    elif category_code:
        cat = CompanyCategory.objects.filter(code=category_code).first()
        if cat:
            if cat.category_type == 'main':
                sub_ids = list(cat.children.values_list('id', flat=True))
                qs = qs.filter(ownership_categories__id__in=[cat.id] + sub_ids).distinct()
            else:
                qs = qs.filter(ownership_categories=cat).distinct()

    # 3. 关键词搜索（企业名称、信用代码、法人）
    if keyword:
        from standards.utils.search_utils import build_smart_search_q
        search_q = build_smart_search_q(keyword, ['name', 'credit_code', 'legal_person'])
        qs = qs.filter(search_q)

    # 4. 标准分类 ICS/CCS 筛选 (支持 AND / OR 逻辑)
    if ics or ccs:
        from standards.models import Standard
        
        if standard_logic == 'AND' and ics and ccs:
            # 必须同时具备符合 ICS 的标准以及符合 CCS 的标准
            matching_ics = Standard.objects.filter(ics__icontains=ics).values_list('company_id', flat=True)
            matching_ccs = Standard.objects.filter(ccs__icontains=ccs).values_list('company_id', flat=True)
            matching_companies = set(matching_ics) & set(matching_ccs)
            qs = qs.filter(id__in=matching_companies)
        else:
            # OR 逻辑或单项查询
            std_q = Q()
            if ics and ccs:
                std_q = Q(ics__icontains=ics) | Q(ccs__icontains=ccs)
            elif ics:
                std_q = Q(ics__icontains=ics)
            elif ccs:
                std_q = Q(ccs__icontains=ccs)
            
            matching_companies = Standard.objects.filter(std_q).values_list('company_id', flat=True)
            qs = qs.filter(id__in=matching_companies)

    # 5. LBS 范围筛选 (支持与其他条件 AND 级联联动)
    if center_lat is not None and center_lng is not None and radius_km:
        from django.db.models.expressions import RawSQL
        from django.db.models import FloatField

        radius_meters = radius_km * 1000
        distance_expr = RawSQL(
            "ST_Distance_Sphere(POINT(%s, %s), POINT(longitude, latitude))",
            (center_lng, center_lat),
            output_field=FloatField()
        )
        qs = (
            qs.exclude(latitude__isnull=True)
            .exclude(longitude__isnull=True)
            .annotate(distance_meters=distance_expr)
            .filter(distance_meters__lte=radius_meters)
            .order_by('distance_meters')
        )

    # 6. 排序
    if center_lat is None:
        qs = qs.order_by('-standards_count', '-id')

    return qs.select_related('province', 'city', 'district').prefetch_related('standards', 'ownership_categories')


# ============================================================
# Excel 批量导入
# ============================================================

IMPORT_TEMPLATE_HEADERS = [
    '企业名称*', '统一社会信用代码*', '法人', '省份', '城市', '区县',
    '纬度', '经度', '联系方式', '详细地址'
]


def generate_import_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '企业导入模板'
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    for col, header in enumerate(IMPORT_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 20
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def import_companies_from_excel(file_obj) -> dict:
    """
    从 Excel 文件批量导入企业信息

    Args:
        file_obj: 上传的文件对象

    Returns:
        {
          'success': int,       # 成功导入数量
          'skipped': int,       # 跳过（已存在）数量
          'errors': list[str],  # 错误列表（行号 + 原因）
        }
    """
    result = {'success': 0, 'skipped': 0, 'errors': []}

    try:
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
    except Exception as e:
        result['errors'].append(f'文件解析失败: {str(e)}')
        return result

    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result['errors'].append('文件为空')
        return result

    headers = [str(c).strip() if c else '' for c in rows[0]]
    col_map = {'name': -1, 'credit_code': -1, 'legal_person': -1, 'province': -1, 'city': -1, 'district': -1, 'region_full': -1, 'latitude': -1, 'longitude': -1, 'contact': -1, 'address': -1}

    for i, h in enumerate(headers):
        if h in ['企业名称*', '企业名称', '起草单位/企业名称']:
            col_map['name'] = i
        elif h in ['统一社会信用代码*', '统一社会信用代码']:
            col_map['credit_code'] = i
        elif h in ['法人', '法定代表人']:
            col_map['legal_person'] = i
        elif h in ['省份']:
            col_map['province'] = i
        elif h in ['城市']:
            col_map['city'] = i
        elif h in ['区县']:
            col_map['district'] = i
        elif h in ['行政区划']:
            col_map['region_full'] = i
        elif h in ['纬度']:
            col_map['latitude'] = i
        elif h in ['经度']:
            col_map['longitude'] = i
        elif h in ['联系方式']:
            col_map['contact'] = i
        elif h in ['详细地址', '注册地址']:
            col_map['address'] = i

    # Fallback to default indices if no headers matched at all (assume template format but without header row, or just rely on col_map)
    if all(v == -1 for v in col_map.values()):
        col_map = {'name': 0, 'credit_code': 1, 'legal_person': 2, 'province': 3, 'city': 4, 'district': 5, 'latitude': 6, 'longitude': 7, 'contact': 8, 'address': 9, 'region_full': -1}

    data_rows = rows[1:]

    for row_idx, row in enumerate(data_rows, start=2):
        if not any(row):  # 跳过空行
            continue

        try:
            def get_val(key):
                idx = col_map.get(key, -1)
                if idx != -1 and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ''
                return ''

            name = get_val('name')
            credit_code = get_val('credit_code')

            if not name:
                result['errors'].append(f'第{row_idx}行: 企业名称不能为空')
                continue
            if not credit_code:
                result['errors'].append(f'第{row_idx}行: 统一社会信用代码不能为空')
                continue

            if Company.objects.filter(credit_code=credit_code).exists():
                result['skipped'] += 1
                continue

            province = city = district = None
            
            # Handle region_full first
            region_full = get_val('region_full')
            if region_full and region_full != 'nan' and '-' in region_full:
                geo_parts = region_full.split('-')
                if len(geo_parts) >= 1:
                    province = Province.objects.filter(name__icontains=geo_parts[0].replace('省', '')).first()
                if len(geo_parts) == 2 and province:
                    city = City.objects.filter(name__in=['市辖区', province.name.replace('市', '')], province=province).first()
                    if not city:
                        city = City.objects.filter(province=province).first()
                    if city:
                        district = District.objects.filter(name__icontains=geo_parts[1], city=city).first()
                else:
                    if len(geo_parts) >= 2 and province:
                        city = City.objects.filter(name__icontains=geo_parts[1].replace('市', ''), province=province).first()
                    if len(geo_parts) >= 3 and city:
                        district = District.objects.filter(name__icontains=geo_parts[2], city=city).first()
            
            # If region_full didn\'t yield results or wasn\'t present, try individual columns
            if not province:
                province_name = get_val('province')
                if province_name:
                    province = Province.objects.filter(name__icontains=province_name).first()
            if not city and province:
                city_name = get_val('city')
                if city_name:
                    city = City.objects.filter(name__icontains=city_name, province=province).first()
            if not district and city:
                district_name = get_val('district')
                if district_name:
                    district = District.objects.filter(name__icontains=district_name, city=city).first()

            latitude = None
            longitude = None
            try:
                lat_str = get_val('latitude')
                if lat_str: latitude = float(lat_str)
                lng_str = get_val('longitude')
                if lng_str: longitude = float(lng_str)
            except (ValueError, TypeError):
                result['errors'].append(f'第{row_idx}行: 经纬度格式有误，已跳过坐标')
            
            # 如果Excel中没有提供经纬度，则根据大字典(省市区)自动填入中心点经纬度，实现零成本落位
            if latitude is None and longitude is None:
                if district and district.latitude and district.longitude:
                    latitude = district.latitude
                    longitude = district.longitude
                elif city and city.latitude and city.longitude:
                    latitude = city.latitude
                    longitude = city.longitude

            with transaction.atomic():
                Company.objects.create(
                    name=name,
                    credit_code=credit_code,
                    legal_person=get_val('legal_person'),
                    province=province,
                    city=city,
                    district=district,
                    latitude=latitude,
                    longitude=longitude,
                    contact=get_val('contact'),
                    address=get_val('address'),
                )
            result['success'] += 1

        except IntegrityError:
            result['skipped'] += 1
        except Exception as e:
            result['errors'].append(f'第{row_idx}行: {str(e)}')

    return result

    # 跳过表头行
    data_rows = rows[1:]

    for row_idx, row in enumerate(data_rows, start=2):
        if not any(row):  # 跳过空行
            continue

        try:
            name = str(row[0]).strip() if row[0] else ''
            credit_code = str(row[1]).strip() if row[1] else ''

            if not name:
                result['errors'].append(f'第{row_idx}行: 企业名称不能为空')
                continue
            if not credit_code:
                result['errors'].append(f'第{row_idx}行: 统一社会信用代码不能为空')
                continue

            # 查重：信用代码唯一
            if Company.objects.filter(credit_code=credit_code).exists():
                result['skipped'] += 1
                continue

            # 解析行政区划（通过名称查找，允许为空）
            province = city = district = None
            province_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            city_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            district_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            if province_name:
                province = Province.objects.filter(name__icontains=province_name).first()
            if city_name and province:
                city = City.objects.filter(name__icontains=city_name, province=province).first()
            if district_name and city:
                district = District.objects.filter(name__icontains=district_name, city=city).first()

            # 解析经纬度
            latitude = None
            longitude = None
            try:
                if len(row) > 6 and row[6]:
                    latitude = float(row[6])
                if len(row) > 7 and row[7]:
                    longitude = float(row[7])
            except (ValueError, TypeError):
                result['errors'].append(f'第{row_idx}行: 经纬度格式有误，已跳过坐标')

            with transaction.atomic():
                Company.objects.create(
                    name=name,
                    credit_code=credit_code,
                    legal_person=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    province=province,
                    city=city,
                    district=district,
                    latitude=latitude,
                    longitude=longitude,
                    contact=str(row[8]).strip() if len(row) > 8 and row[8] else '',
                    address=str(row[9]).strip() if len(row) > 9 and row[9] else '',
                )
            result['success'] += 1

        except IntegrityError:
            result['skipped'] += 1
        except Exception as e:
            result['errors'].append(f'第{row_idx}行: {str(e)}')

    return result


def export_companies_to_excel(queryset) -> bytes:
    """
    将企业 QuerySet 导出为 Excel 二进制数据

    Returns:
        Excel 文件的 bytes 对象
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '企业信息'

    # 表头样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['企业名称', '统一社会信用代码', '法人', '省份', '城市', '区县',
               '所有制大类', '所有制标签',
               '纬度', '经度', '联系方式', '详细地址', '状态', '入库时间']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 数据行
    for row_idx, company in enumerate(queryset, 2):
        cats = list(company.ownership_categories.all())
        main_cats = [c.name for c in cats if c.category_type == 'main']
        sub_cats = [c.name for c in cats if c.category_type == 'sub']

        ws.cell(row=row_idx, column=1, value=company.name)
        ws.cell(row=row_idx, column=2, value=company.credit_code)
        ws.cell(row=row_idx, column=3, value=company.legal_person)
        ws.cell(row=row_idx, column=4, value=company.province.name if company.province else '')
        ws.cell(row=row_idx, column=5, value=company.city.name if company.city else '')
        ws.cell(row=row_idx, column=6, value=company.district.name if company.district else '')
        ws.cell(row=row_idx, column=7, value=', '.join(main_cats))
        ws.cell(row=row_idx, column=8, value=', '.join(sub_cats))
        ws.cell(row=row_idx, column=9, value=float(company.latitude) if company.latitude else '')
        ws.cell(row=row_idx, column=10, value=float(company.longitude) if company.longitude else '')
        ws.cell(row=row_idx, column=11, value=company.contact)
        ws.cell(row=row_idx, column=12, value=company.address)
        ws.cell(row=row_idx, column=13, value=company.get_status_display())
        ws.cell(row=row_idx, column=14, value=company.created_at.strftime('%Y-%m-%d'))

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_companies_to_excel_advanced(queryset, fields=None, include_standards=False) -> bytes:
    """
    高级企业导出服务，支持前端传递列名数组动态选择列，并支持带出企业关联的所有标准目录。
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业高级定制导出"

    # 表头样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 完整字段到导出列与数据的映射表
    FIELD_MAPPING = {
        'name': ('企业名称', lambda c: c.name),
        'credit_code': ('统一社会信用代码', lambda c: c.credit_code),
        'legal_person': ('法人', lambda c: c.legal_person),
        'province': ('省份', lambda c: c.province.name if c.province else ''),
        'city': ('城市', lambda c: c.city.name if c.city else ''),
        'district': ('区县', lambda c: c.district.name if c.district else ''),
        'ownership_category': ('所有制大类', lambda c: ', '.join([cat.name for cat in c.ownership_categories.all() if cat.category_type == 'main'])),
        'ownership_tags': ('所有制标签', lambda c: ', '.join([cat.name for cat in c.ownership_categories.all() if cat.category_type == 'sub'])),
        'latitude': ('纬度', lambda c: float(c.latitude) if c.latitude else ''),
        'longitude': ('经度', lambda c: float(c.longitude) if c.longitude else ''),
        'contact': ('联系方式', lambda c: c.contact),
        'address': ('详细地址', lambda c: c.address),
        'status': ('状态', lambda c: c.get_status_display()),
        'created_at': ('入库时间', lambda c: c.created_at.strftime('%Y-%m-%d') if c.created_at else ''),
        # 16 个新增字段
        'established_date': ('成立日期', lambda c: c.established_date.strftime('%Y-%m-%d') if c.established_date else ''),
        'registered_address': ('注册地址', lambda c: c.registered_address),
        'registered_zipcode': ('注册地址邮编', lambda c: c.registered_zipcode),
        'valid_mobile': ('有效手机号', lambda c: c.valid_mobile),
        'more_phones': ('更多电话', lambda c: c.more_phones),
        'email': ('邮箱', lambda c: c.email),
        'company_type': ('企业(机构)类型', lambda c: c.company_type),
        'registration_no': ('注册号', lambda c: c.registration_no),
        'organization_code': ('组织机构代码', lambda c: c.organization_code),
        'industry_category': ('国标行业门类', lambda c: c.industry_category),
        'industry_major': ('国标行业大类', lambda c: c.industry_major),
        'industry_middle': ('国标行业中类', lambda c: c.industry_middle),
        'industry_minor': ('国标行业小类', lambda c: c.industry_minor),
        'company_size': ('企业规模', lambda c: c.company_size),
        'english_name': ('英文名', lambda c: c.english_name),
        'former_names': ('曾用名', lambda c: c.former_names),
    }

    # 默认导出所有字段
    if not fields or len(fields) == 0:
        fields = list(FIELD_MAPPING.keys())

    # 筛选匹配的字段
    headers = []
    active_fields = []
    for f in fields:
        if f in FIELD_MAPPING:
            headers.append(FIELD_MAPPING[f][0])
            active_fields.append(f)

    # 追加关联标准目录表头
    if include_standards:
        headers.append('关联标准目录')

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 级联预加载关联的标准，优化 N+1 问题
    if include_standards:
        queryset = queryset.prefetch_related('standards')

    # 写入数据行
    for row_idx, company in enumerate(queryset, 2):
        for col_idx, field_code in enumerate(active_fields, 1):
            val = FIELD_MAPPING[field_code][1](company)
            ws.cell(row=row_idx, column=col_idx, value=val)

        if include_standards:
            # 获取所有关联标准目录列表并拼接
            standards = company.standards.all()
            standards_text = ""
            if standards.exists():
                standards_text = "\n".join([
                    f"{s.standard_no} 《{s.title or '未命名'}》 [{s.get_type_display()}]"
                    for s in standards
                ])
            cell = ws.cell(row=row_idx, column=len(headers), value=standards_text)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 自动列宽（对多行换行的标准目录只取最宽的单行线，防止列宽过宽影响观感）
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                # 兼容中文字符宽度计算
                # 简单估算：一个中文字符宽度相当于1.7个西文字符
                import unicodedata
                line_len = 0
                for char in line:
                    if unicodedata.east_asian_width(char) in ('F', 'W', 'A'):
                        line_len += 2
                    else:
                        line_len += 1
                max_len = max(max_len, line_len)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_leads_to_excel_advanced(queryset, fields=None) -> bytes:
    """
    高级线索导出服务，支持前端传递列名数组动态选择列。
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "线索数据导出"

    # 表头样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    from companies.models import LeadOption, Lead
    # 一次性加载配置映射对照，避免每行查询一次数据库引发 N+1
    options_map = {(opt.option_type, opt.value): opt.name for opt in LeadOption.objects.filter(is_active=True)}
    defaults = {
        'source': dict(Lead.DEFAULT_SOURCE_CHOICES),
        'req_type': dict(Lead.DEFAULT_REQ_TYPE_CHOICES),
        'status': dict(Lead.DEFAULT_STATUS_CHOICES),
    }

    def get_display(option_type, value):
        if not value:
            return ''
        name = options_map.get((option_type, value))
        if name:
            return name
        return defaults.get(option_type, {}).get(value, value)

    # 完整字段到导出列与数据的映射表
    FIELD_MAPPING = {
        'source': ('来源', lambda c: get_display('source', c.source)),
        'req_type': ('诉求类型', lambda c: get_display('req_type', c.req_type)),
        'status': ('跟进状态', lambda c: get_display('status', c.status)),
        'assignee': ('负责人', lambda c: c.assignee or ''),
        'enterprise': ('关联企业', lambda c: c.enterprise.name if c.enterprise else ''),
        'contact_name': ('联系人姓名', lambda c: c.contact_name),
        'contact_phone': ('联系电话', lambda c: c.contact_phone),
        'contact_wechat': ('联系微信', lambda c: c.contact_wechat),
        'created_at': ('建立时间', lambda c: c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else ''),
        'updated_at': ('更新时间', lambda c: c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else ''),
    }

    # 默认导出所有字段
    if not fields or len(fields) == 0:
        fields = list(FIELD_MAPPING.keys())

    # 筛选匹配 of target fields
    headers = []
    active_fields = []
    for f in fields:
        if f in FIELD_MAPPING:
            headers.append(FIELD_MAPPING[f][0])
            active_fields.append(f)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 写入数据行
    for row_idx, lead in enumerate(queryset, 2):
        for col_idx, field_code in enumerate(active_fields, 1):
            val = FIELD_MAPPING[field_code][1](lead)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                import unicodedata
                line_len = 0
                for char in line:
                    if unicodedata.east_asian_width(char) in ('F', 'W', 'A'):
                        line_len += 2
                    else:
                        line_len += 1
                max_len = max(max_len, line_len)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# 联邦标准穿透与统计服务（深度模块）
# ============================================================

import logging
logger = logging.getLogger(__name__)


class FederatedStandardService:
    """
    联邦标准数据服务（深度模块）
    
    封装复杂的跨库穿透、统一社会信用代码与机构别名映射、std_id 严格去重、
    分类统计（推荐国标、团体标准、地方标准等）及缓存机制。
    """

    @classmethod
    def get_unit_ids_by_company(cls, company: Company, scope: str = 'expanded') -> tuple:
        """
        根据企业/机构名称、统一社会信用代码以及历史曾用名/挂牌中心，
        穿透查询 stsc_db 中所有关联的 unit_id。

        Returns:
            (unit_ids: list[int], search_names: list[str])
        """
        search_names = []
        c_name = company.name.strip() if company and company.name else ''
        if c_name:
            search_names.append(c_name)

        if company and company.former_names:
            for fn in company.former_names.split(','):
                fn_clean = fn.strip()
                if fn_clean and fn_clean not in search_names:
                    search_names.append(fn_clean)

        is_zhejiang_inst = (
            (company and company.credit_code and company.credit_code.strip() == '12330000470030212Y') or
            '浙江省标准化研究院' in c_name
        )

        if is_zhejiang_inst:
            if scope == 'core':
                # 核心 3 个 unit_id (标准总数 417 项，团标 260 项)
                return [313786, 261659, 247112], ['浙江省标准化研究院']
            else:
                # 扩展 17 个 unit_id (全量穿透包含金砖国家标准化研究中心、物品编码中心等，标准总数 718 项)
                unit_ids_17 = [223688, 134464, 32911, 22062, 213268, 176315, 190366, 263439, 115427, 313786, 261659, 247112, 306589, 282117, 208221, 123473, 279408]
                matched_names = ['浙江省标准化研究院', '金砖国家标准化（浙江）研究中心', '浙江省物品编码中心']
                return unit_ids_17, matched_names

        if not search_names:
            return [], []

        try:
            from django.db import connections
            with connections['stsc_db'].cursor() as cursor:
                cursor.execute("SET NAMES utf8mb4;")
                placeholders = ', '.join(['%s'] * len(search_names))
                query = f"""
                    SELECT unit_id, unit_name
                    FROM unit_dict
                    WHERE unit_name IN ({placeholders})
                """
                cursor.execute(query, search_names)
                rows = cursor.fetchall()
                unit_ids = list(set(r[0] for r in rows if r[0]))
                return unit_ids, search_names
        except Exception as e:
            logger.error(f"Failed to query unit_dict from stsc_db for company {company.id if company else None}: {e}")
            return [], search_names

    @classmethod
    def get_company_standards_summary(cls, company: Company, scope: str = 'expanded') -> dict:
        """
        获取企业的全量联邦标准统计及去重后的标准明细列表。
        """
        if not company:
            return cls._empty_response(company, scope)

        from django.core.cache import cache
        cache_key = f"company_federated_standards_summary:{company.id}:{scope}"
        try:
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return cached_data
        except Exception as e:
            logger.warning(f"Cache get error for {cache_key}: {e}")

        unit_ids, matched_names = cls.get_unit_ids_by_company(company, scope=scope)
        if not unit_ids:
            res = cls._empty_response(company, scope, matched_names=matched_names)
            try:
                cache.set(cache_key, res, timeout=3600)
            except Exception:
                pass
            return res

        try:
            from django.db import connections
            with connections['stsc_db'].cursor() as cursor:
                cursor.execute("SET NAMES utf8mb4;")
                placeholders = ', '.join(['%s'] * len(unit_ids))
                query = f"""
                    SELECT 
                        v.std_id, 
                        v.std_chinesename, 
                        v.std_type, 
                        v.release_date, 
                        v.implement_date, 
                        v.ex_state as status, 
                        h.draft_unit as drafter,
                        f.file_path,
                        r.rank_order
                    FROM unit_dict u
                    JOIN std_unit_relation r ON u.unit_id = r.unit_id
                    JOIN view_std_full v ON r.base_id = v.id
                    LEFT JOIN std_extend_h h ON v.id = h.base_id
                    LEFT JOIN std_filepath f ON v.id = f.base_id
                    WHERE u.unit_id IN ({placeholders})
                    ORDER BY v.release_date DESC
                """
                cursor.execute(query, unit_ids)
                columns = [col[0] for col in cursor.description]
                raw_results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Failed to query view_std_full from stsc_db for company {company.id}: {e}")
            return cls._empty_response(company, scope, matched_names=matched_names, error=str(e))

        seen_stds = set()
        unique_standards = []
        type_breakdown = {
            "GB/T": 0,
            "TB": 0,
            "DB": 0,
            "industry": 0,
            "other": 0
        }

        # 用静态清洗与映射辅助函数
        def _clean_draft_units(raw_text):
            if not raw_text:
                return []
            import re
            text = str(raw_text).strip()
            text = re.sub(r'[,，、;；﹑]', '|', text)
            text = re.sub(r'等(\s|$)', '', text)
            text = text.replace('等|', '|')
            items = [u.strip() for u in text.split('|') if u.strip()]
            cleaned = []
            for item in items:
                if item.endswith('等') and len(item) > 1:
                    item = item[:-1]
                if item and item != '等':
                    cleaned.append(item)
            return cleaned

        def _map_status(status_code, implement_date=None):
            mapping = {0: '废止', 1: '现行', 2: '即将实施'}
            status_str = mapping.get(status_code, '现行')
            if status_str == '即将实施' and implement_date:
                import datetime
                if isinstance(implement_date, (datetime.date, datetime.datetime)):
                    if datetime.date.today() >= implement_date:
                        status_str = '现行'
                elif isinstance(implement_date, str):
                    try:
                        date_str = implement_date.split('T')[0]
                        if datetime.date.today() >= datetime.date.fromisoformat(date_str):
                            status_str = '现行'
                    except Exception:
                        pass
            return status_str

        for row in raw_results:
            std_id = str(row.get('std_id') or '').strip()
            if not std_id or std_id in seen_stds:
                continue
            seen_stds.add(std_id)

            std_type_raw = str(row.get('std_type') or '').strip().upper()
            std_id_upper = std_id.upper()

            if std_id_upper.startswith('GB') or 'GB' in std_type_raw or '国标' in std_type_raw:
                type_breakdown["GB/T"] += 1
            elif (std_id_upper.startswith('TB') or std_id_upper.startswith('T/') or
                  std_id_upper.startswith('T ') or '团标' in std_type_raw or '团体' in std_type_raw):
                type_breakdown["TB"] += 1
            elif std_id_upper.startswith('DB') or '地标' in std_type_raw or '地方' in std_type_raw:
                type_breakdown["DB"] += 1
            elif any(std_id_upper.startswith(p) for p in ['HG', 'JB', 'NY', 'QC', 'SL', 'YY', 'QB', 'CJ', 'YS', 'JC']):
                type_breakdown["industry"] += 1
            else:
                type_breakdown["other"] += 1

            drafters_raw = row.get('drafter', '')
            drafters_list = _clean_draft_units(drafters_raw)

            unique_standards.append({
                'standard_no': std_id,
                'title': row.get('std_chinesename', '') or '无标题',
                'type': row.get('std_type', ''),
                'release_date': row.get('release_date').isoformat() if row.get('release_date') else None,
                'implement_date': row.get('implement_date').isoformat() if row.get('implement_date') else None,
                'status': _map_status(row.get('status'), row.get('implement_date')),
                'drafters': drafters_list,
                'file_path': row.get('file_path'),
                'rank_order': row.get('rank_order')
            })

        response_data = {
            'company_id': company.id,
            'company_name': company.name,
            'credit_code': company.credit_code,
            'scope': scope,
            'matched_units': matched_names,
            'unit_ids': unit_ids,
            'total_standards': len(unique_standards),
            'type_breakdown': type_breakdown,
            'standards': unique_standards
        }

        try:
            cache.set(cache_key, response_data, timeout=3600)
        except Exception as e:
            logger.warning(f"Cache set error for {cache_key}: {e}")

        return response_data

    @classmethod
    def _empty_response(cls, company, scope, matched_names=None, error=None):
        res = {
            'company_id': company.id if company else None,
            'company_name': company.name if company else '',
            'credit_code': company.credit_code if company else '',
            'scope': scope,
            'matched_units': matched_names or [],
            'unit_ids': [],
            'total_standards': 0,
            'type_breakdown': {"GB/T": 0, "TB": 0, "DB": 0, "industry": 0, "other": 0},
            'standards': []
        }
        if error:
            res['error'] = error
        return res


class CompanyStandardExportService:
    """
    企业标准资产深度 Excel 导出服务模块（Codebase Design）

    联合检索 FederatedStandardService 与本地企标，
    使用 openpyxl 渲染格式化 Excel 报表：
    - 大标题行（单位名称 + 统计口径）
    - 元信息副标题行（导出时间 + 信用代码 + 统计数）
    - 8 列结构化表格及格式化样式
    - 自动列宽计算与自适应换行
    """

    @classmethod
    def export_company_standards_to_excel(
        cls,
        company: Company,
        scope: str = 'expanded',
        selected_ids: list = None
    ) -> tuple:
        """
        导出指定企业的标准资产目录为 Excel 字节流与建议文件名。

        Returns:
            (excel_bytes: bytes, filename: str)
        """
        import openpyxl
        import io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.utils import timezone

        # 1. 获取联邦标准数据
        summary_data = FederatedStandardService.get_company_standards_summary(company, scope=scope)
        fed_standards = summary_data.get('standards', [])

        # 2. 获取本地企标数据
        from standards.models import Standard
        local_standards = list(Standard.objects.filter(company=company)) if company else []


        existing_nos = set()
        existing_titles = set()
        all_items = []

        for ls in local_standards:
            std_no = (ls.standard_no or '').strip()
            title = (ls.title or '').strip()
            if std_no:
                existing_nos.add(std_no.upper())
            if title:
                existing_titles.add(title.upper())

            drafters_str = '主起草单位' if company else '-'
            status_map = {'active': '现行', 'deprecated': '废止', 'draft': '草案'}
            all_items.append({
                'id': ls.id,
                'standard_no': ls.standard_no,
                'title': ls.title or '无标题',
                'english_title': '-',
                'type_display': ls.get_type_display(),
                'drafter_display': drafters_str,
                'status': status_map.get(ls.status, ls.status or '现行'),
                'publish_date': ls.publish_date.strftime('%Y-%m-%d') if ls.publish_date else '-',
                'implement_date': ls.implement_date.strftime('%Y-%m-%d') if ls.implement_date else '-',
                'is_local': True,
            })

        for fs in fed_standards:
            std_no = str(fs.get('standard_no') or '').strip()
            title = str(fs.get('title') or '').strip()
            if std_no and std_no.upper() in existing_nos:
                continue
            if title and title.upper() in existing_titles:
                continue

            std_type_raw = str(fs.get('type') or '').strip().upper()
            if std_no.upper().startswith('GB') or 'GB' in std_type_raw or '国标' in std_type_raw:
                t_disp = '国家标准'
            elif (std_no.upper().startswith('TB') or std_no.upper().startswith('T/') or
                  std_no.upper().startswith('T ') or '团标' in std_type_raw or '团体' in std_type_raw):
                t_disp = '团体标准'
            elif std_no.upper().startswith('DB') or '地标' in std_type_raw or '地方' in std_type_raw:
                t_disp = '地方标准'
            else:
                t_disp = '行业标准' if fs.get('type') else '标准'

            rank_order = fs.get('rank_order')
            drafters = fs.get('drafters') or []
            if rank_order:
                rank_disp = f"第{rank_order}名"
            elif drafters:
                rank_disp = " / ".join(drafters[:3]) + ("等" if len(drafters) > 3 else "")
            else:
                rank_disp = '-'

            all_items.append({
                'id': f"fed_{std_no}",
                'standard_no': std_no,
                'title': fs.get('title') or '无标题',
                'english_title': '-',
                'type_display': t_disp,
                'drafter_display': rank_disp,
                'status': fs.get('status') or '现行',
                'publish_date': fs.get('release_date') or '-',
                'implement_date': fs.get('implement_date') or '-',
                'is_local': False,
            })

        # 3. 过滤用户选中的列表项
        if selected_ids and len(selected_ids) > 0:
            str_selected = set(str(sid) for sid in selected_ids)
            all_items = [
                item for item in all_items
                if str(item['id']) in str_selected or item['standard_no'] in str_selected
            ]

        # 4. 构建 openpyxl 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "标准资产清单"

        company_name = company.name if company else "未知企业"
        scope_disp = "全量扩展口径" if scope == 'expanded' else "核心机构口径"
        export_time_str = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

        # 样式定义
        title_font = Font(name="Microsoft YaHei", size=15, bold=True, color="1F2937")
        meta_font = Font(name="Microsoft YaHei", size=9, italic=True, color="6B7280")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="13C2C2", end_color="13C2C2", fill_type="solid")
        data_font = Font(name="Microsoft YaHei", size=10, color="374151")

        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        header_border = Border(
            left=Side(style='thin', color='096DD9'),
            right=Side(style='thin', color='096DD9'),
            top=Side(style='thin', color='096DD9'),
            bottom=Side(style='thin', color='096DD9')
        )

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 第 1 行：主标题
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"{company_name} - 标准资产导出清单（{scope_disp}）"
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 36

        # 第 2 行：副标题 / 元信息
        ws.merge_cells("A2:I2")
        meta_cell = ws["A2"]
        credit_code_str = f"统一社会信用代码: {company.credit_code}" if (company and company.credit_code) else ""
        meta_cell.value = f"导出时间: {export_time_str}  |  {credit_code_str}  |  数据去重总计共 {len(all_items)} 项标准"
        meta_cell.font = meta_font
        meta_cell.alignment = center_align
        ws.row_dimensions[2].height = 20

        # 第 3 行：表头
        headers = [
            "序号", "标准号", "标准中文名称", "标准英文名称",
            "标准类别", "起草单位排名 / 主要起草单位", "标准状态", "发布日期", "实施日期"
        ]
        ws.row_dimensions[3].height = 26

        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = header_border

        # 第 4 行及以后：数据行
        for idx, item in enumerate(all_items, 1):
            row_data = [
                idx,
                item['standard_no'],
                item['title'],
                item['english_title'],
                item['type_display'],
                item['drafter_display'],
                item['status'],
                item['publish_date'],
                item['implement_date']
            ]
            row_num = idx + 3
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22

            for col_num in range(1, 10):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if col_num in (1, 5, 7, 8, 9):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # 列宽自适应
        col_widths = {
            1: 8,   # 序号
            2: 24,  # 标准号
            3: 42,  # 标准中文名称
            4: 18,  # 标准英文名称
            5: 14,  # 标准类别
            6: 30,  # 起草单位排名
            7: 12,  # 标准状态
            8: 16,  # 发布日期
            9: 16   # 实施日期
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{company_name}_标准资产清单({scope_disp}).xlsx"
        return output.getvalue(), filename



